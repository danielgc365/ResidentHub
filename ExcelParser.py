################################################################################
# ExcelParser takes a GM workbook (SBR or main) and obtains the necessary data
# to create a calibration
#
#
# Usage for SDM Veoneer Residents ONLY
# Created by Daniel Gomez
# Date 10/27/2021
##################################################

# Imports
from openpyxl import Workbook, load_workbook
import os, os.path
import win32com.client as win32

#############################################################
# Class GM workbook
# Takes the GM Workbook excel file and parse necessary data
# Input: Path to GM Workbook
#############################################################
class GMWorkbook:
    def __init__(self, path):
        self.path = path
        self.name = self.path[self.path.rfind('\\') + 1:]
        self.wb = load_workbook(self.path, keep_vba=True, read_only=True)
        pass

    def open_sheet(self, sheet="Release Form"):
        self.Sheet = self.wb[sheet]

    def get_sheet_list(self):
        for sheet in self.wb:
            print(sheet.title)

    def parse(self):
        try:
            self.CalPN = self.Sheet['E21'].value
            self.UtilityPN = self.Sheet['E19'].value
            self.AppPN = self.Sheet['E25'].value
            self.CalAlphaCode = self.Sheet['F21'].value + self.Sheet['G21'].value
            self.AppAlphaCode = self.Sheet['F25'].value + self.Sheet['G25'].value
            self.ModelYear = self.Sheet['E4'].value
            self.EMPN = self.Sheet['E13'].value
            self.SDMType = self.Sheet['C13'].value
        except AttributeError:
            print("No sheet selected")

    def run_macro(self, MacroName="CopyDPID_DIDs"):
        xl = win32.Dispatch('Excel.Application')
        #xl.Application.visible = False  # change to True if you are desired to make Excel visible

        try:
            wb = xl.Workbooks.Open(os.path.abspath(self.path))
            xl.Application.run(self.name + "!Instructions." + MacroName)  # TODO Remove "instructions."

        except Exception as ex:
            template = "An exception of type {0} occurred. Arguments:\n{1!r}"
            message = template.format(type(ex).__name__, ex.args)
            print(message)

        xl.Application.Quit()
        del xl


#############################################################
# Class GM SBR workbook
# Inherits from the GM Workbook class
# it parses the data from the SBR excel file
# Input: Path to GM SBR Workbook
#############################################################
class SBRWorkbook(GMWorkbook):
    def __init__(self, path):
        super().__init__(path)

    def parse(self):
        try:
            self.CalPN = self.Sheet['D34'].value
            self.CalAlphaCode = self.Sheet['E34'].value + self.Sheet['F34'].value
        except AttributeError:
            print("Error")


#############################################################
# Class PRN Tool
# Inherits from the GM Workbook class
# it opens the Veoneer PRN conversion tool, sets the
# correct files, and runs the macros to create the VNR PRN
# Input: Path to appropriate PRN tool
#############################################################
class PRNTool(GMWorkbook):
    def __init__(self, path, MainWbPath, SBRWbPath):
        super().__init__(path)
        self.MainWbPath = MainWbPath
        self.SBRWbPath = SBRWbPath


test_main = GMWorkbook(r'C:\Users\daniel.gomez\PycharmProjects\ResidentHub\85545780_A_B_20210629.xlsm')
test_SBR = SBRWorkbook(r'C:\Users\daniel.gomez\PycharmProjects\ResidentHub\85545781_A_A_20210601.xlsm')
# test_main.open_sheet()
# test_main.parse()
# test_main.run_macro()


test_prn_tool = PRNTool(r'C:\Users\daniel.gomez\PycharmProjects\ResidentHub\GM PRN '
                        r'Tool\SDM50_GM_PRN_Conversion_Tool_23.10.159_V1.xlsm', test_main.path, test_SBR.path)
test_prn_tool.open_sheet("Instructions")
test_prn_tool.Sheet.CommandButton1_Click


# test_prn_tool.run_macro("ClearVars_Click")

# TODO Run macro with dialog window
# TODO Make self.name the path without the initial stuff
# TODO Do entire run of excel stuff